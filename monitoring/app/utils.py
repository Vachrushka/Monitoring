from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, NamedStyle, Alignment
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
from io import BytesIO


def create_style(book, font_color, fill_color, name):
    black_fill = PatternFill(fill_type='solid', start_color=fill_color)
    white_font = Font(color=font_color)
    black_style = NamedStyle(name)
    black_style.fill = black_fill
    black_style.font = white_font
    black_style.alignment = Alignment(
        horizontal='general',
        vertical='bottom',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=True,
        indent=0)
    book.add_named_style(black_style)
    return black_style


def get_excel_io(student, start_date, end_date, data):
    book = Workbook()
    book.active = 0
    sh = book.active
    sh.title = str(student)
    black_style = create_style(book, 'ffffff', '000000', 'bl_s')
    grey_style = create_style(book, 'ffffff', '595959', 'gl_s')
    normal_grey_style = create_style(book, 'ffffff', '787878', 'ng_s')
    lite_grey_style = create_style(book, 'ffffff', 'A5A5A5', 'lg_s')
    lite_style = create_style(book, '000000', 'd8d8d8', 'l_s')
    sh.cell(2, 1, "ФИО").style = black_style
    sh.cell(2, 2, str(student)).style = lite_style
    sh.cell(3, 1, "Направление").style = black_style
    sh.cell(3, 2, str(student.departament)).style = lite_style
    sh.cell(4, 1, "Курс").style = black_style
    sh.cell(4, 2, str(student.course)).style = lite_style
    sh.cell(5, 1, "Дата (от)").style = black_style
    sh.cell(5, 2, start_date.strftime("%d-%m-%Y")).style = lite_style
    sh.cell(6, 1, "Дата (до)").style = black_style
    sh.cell(6, 2, end_date.strftime("%d-%m-%Y")).style = lite_style
    sh.cell(2, 3, "Категория").style = black_style
    sh.cell(2, 4, "Упражнение").style = black_style
    sh.cell(2, 5, "Форма").style = black_style

    start_row = 3
    column = 3
    row_i = start_row
    ind_max = 0
    add_grey_line = []
    add_normal_line = []
    for category, ex in data.items():
        sh.cell(row_i, column, str(category)).style = grey_style
        add_normal_line.append([row_i + 1])
        for e, gr in ex.items():
            sh.cell(row_i, column + 1, e.name).style = grey_style
            form_types = {}
            for g in sorted(gr, key=lambda x: x.datetime):
                if g.rate or g.absence_reason:
                    if g.uniform_type in form_types:
                        form_types[g.uniform_type].append(g)
                        continue
                    form_types[g.uniform_type] = []
                    form_types[g.uniform_type].append(g)
            gr_ind = 0
            for f_t, grade_data in form_types.items():
                sh.cell(row_i, column + 2, f_t.name).style = normal_grey_style
                sh.cell(row_i, column + 3, "Дата").style = lite_style
                sh.cell(row_i + 1, column + 3, "Результат").style = lite_style
                skipped = 0
                ind = 0
                for ind, g in enumerate(grade_data):
                    if g.rate or g.absence_reason:
                        sh.cell(row_i, column + 4 + ind - skipped, g.datetime.strftime("%d/%m/%Y")).style = lite_style
                        sh.cell(row_i + 1, column + 4 + ind - skipped, g.rate if g.rate else str(g.absence_reason))
                    else:
                        skipped += 1
                ind_max = column + 5 + ind - skipped if column + 5 + ind - skipped > ind_max else ind_max
                if gr_ind == len(form_types) - 1:
                    add_grey_line.append((row_i + 2, 4))
                    row_i += 3
                else:
                    row_i += 2
                gr_ind += 1

            if not form_types:
                sh.cell(row_i, column + 2, "Записи отсутствуют").style = normal_grey_style
                row_i += 2
        add_grey_line.append((row_i - 1, 3))
        add_normal_line[-1].append(row_i - 1)
    for i in range(6, ind_max):
        sh.cell(2, i, "").style = black_style
    for row, start in add_grey_line:
        for i in range(start, ind_max):
            sh.cell(row, i, "").style = grey_style
    for row_s, row_e in add_normal_line:
        for i in range(row_s, row_e):
            sh.cell(i, 3, "").style = lite_grey_style

    widths = {'A': 69,
              'B': 150,
              'C': 73,
              'D': 200,
              'E': 100}
    for key, v in widths.items():
        sh.column_dimensions[key].width = 0.17*v

    # data = Reference(sh, min_col=7, min_row=3, max_col=15, max_row=4)
    # c2 = LineChart()
    # c2.title = "Date Axis"
    # c2.style = 12
    # c2.y_axis.title = "Size"
    # c2.y_axis.crossAx = 500
    # c2.x_axis = DateAxis(crossAx=100)
    # c2.x_axis.number_format = 'd-mmm'
    # c2.x_axis.majorTimeUnit = "days"
    # c2.x_axis.title = "Date"
    #
    # c2.add_data(data, titles_from_data=True)
    # # dates = Reference(sh, min_col=7, min_row=3, max_col=15)
    # # c2.set_categories(dates)
    #
    # sh.add_chart(c2, "A61")

    buffer = BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def calculate_points(result, min_r, max_r, koef_min, koef_top, reverse):
    if (not reverse and min_r > max_r) or (reverse and min_r < max_r):
        max_r, min_r = min_r, max_r
    if min_r == max_r:
        return 0
    price = 100 / (max_r - min_r)
    shift = 0
    if (not reverse and result > max_r) or (reverse and result < max_r):
        shift = result - max_r
        result = max_r

    rating = (result - min_r) / (max_r - min_r) * 10 * koef_min + shift * price * koef_top
    return max(round(rating, 2), 0)
